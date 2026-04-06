"""
Multi-sheet Excel workbook generator — separate sheets per conversion stage.
"""
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from cost_engine import calc_alloy_cost, calc_conversion_costs

HF = Font(name='Arial', bold=True, size=10, color='FFFFFF')
HFL = PatternFill('solid', fgColor='2F5496')
DF = Font(name='Arial', size=10); BF = Font(name='Arial', bold=True, size=10)
RF = Font(name='Arial', bold=True, size=10, color='C00000')
GF = Font(name='Arial', size=10, color='008000')
SF = Font(name='Arial', size=9, italic=True, color='666666')
TB = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
TODAY_FILL = PatternFill('solid', fgColor='FFFF00')
AC = {'AA2040':'C00000','AA2050':'2F5496','AA2099':'008000','AA2618':'7030A0','AA7140':'BF6900'}
AF = {k: PatternFill('solid', fgColor=v) for k,v in {'AA2040':'FCE4EC','AA2050':'E3F2FD','AA2099':'E8F5E9','AA2618':'F3E5F5','AA7140':'FFF3E0'}.items()}
SH = {'raw':PatternFill('solid',fgColor='808080'),'billet':PatternFill('solid',fgColor='4BACC6'),'ext':PatternFill('solid',fgColor='70AD47')}
SFill = {'raw':PatternFill('solid',fgColor='F2F2F2'),'billet':PatternFill('solid',fgColor='DAEEF3'),'ext':PatternFill('solid',fgColor='E2EFDA')}

def _hp(row):
    return {k: row[k.lower() if k != 'Ag_oz' else 'ag_oz'] for k in ['Al','Cu','Ag_oz','Zn','Ni','Li','Mg','Mn','Ti','Zr','Fe','Si']}

def _stage_sheet(wb, title, main, sub, df, alloys, ak, rb, rt, stage, rfill, extra=False):
    N=len(df); ws=wb.create_sheet(title); nc=8 if extra else 6
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=nc); ws['A1'].value=main; ws['A1'].font=Font(name='Arial',bold=True,size=14,color='2F5496')
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=nc); ws['A2'].value=sub; ws['A2'].font=SF
    hdrs=['Date']+[alloys[k]['name'] for k in ak]+(['Spread','Ag $/oz'] if extra else [])
    for c,h in enumerate(hdrs,1):
        cl=ws.cell(row=3,column=c,value=h); cl.font=HF; cl.fill=HFL; cl.alignment=Alignment(horizontal='center',wrap_text=True); cl.border=TB
        if 2<=c<=6: cl.fill=PatternFill('solid',fgColor=AC[ak[c-2]])
    for i in range(N):
        r=df.iloc[i]; rn=4+i; last=(i==N-1); lbl=r['date']+(' ← LATEST' if last else '')
        cl=ws.cell(row=rn,column=1,value=lbl); cl.font=RF if last else DF; cl.border=TB
        if last: cl.fill=TODAY_FILL
        hp=_hp(r); costs={}
        for j,key in enumerate(ak):
            raw,_,_=calc_alloy_cost(alloys[key]['comp'],hp); bil,ext=calc_conversion_costs(raw,rb,rt)
            val={'raw':raw,'billet':bil,'ext':ext}[stage]; costs[key]=val
            cl=ws.cell(row=rn,column=2+j,value=round(val,2)); cl.font=Font(name='Arial',bold=True,size=10,color=AC[key])
            cl.fill=TODAY_FILL if last else rfill; cl.border=TB; cl.number_format='0.00'
        if extra:
            sp=max(costs.values())-min(costs.values())
            cl=ws.cell(row=rn,column=7,value=round(sp,2)); cl.font=BF; cl.border=TB; cl.number_format='0.00'
            if last: cl.fill=TODAY_FILL
            cl=ws.cell(row=rn,column=8,value=r['ag_oz']); cl.font=Font(name='Arial',size=10,color='0000FF'); cl.border=TB; cl.number_format='0.00'
            if last: cl.fill=TODAY_FILL
    ws.column_dimensions['A'].width=20
    for c in range(2,nc+1): ws.column_dimensions[get_column_letter(c)].width=12

def generate_excel(df_hist, alloys, conversion):
    wb=openpyxl.Workbook(); ak=list(alloys.keys()); N=len(df_hist)
    rb=conversion['r_billet']; re_=conversion.get('r_extrusion',1.3); rt=conversion['r_total']
    # Sheet 1: Compositions
    ws=wb.active; ws.title="Alloy Compositions"
    ws.merge_cells('A1:N1'); ws['A1'].value='Alloy Composition Comparison (wt%)'; ws['A1'].font=Font(name='Arial',bold=True,size=14,color='2F5496')
    for c,h in enumerate(['Alloy','Spec','Al','Cu','Zn','Mg','Mn','Ag','Li','Ni','Fe','Si','Zr','Ti'],1):
        cl=ws.cell(row=3,column=c,value=h); cl.font=HF; cl.fill=HFL; cl.alignment=Alignment(horizontal='center'); cl.border=TB
    for i,(key,a) in enumerate(alloys.items()):
        r=4+i; ws.cell(row=r,column=1,value=a['name']).font=Font(name='Arial',bold=True,size=10,color=AC[key]); ws.cell(row=r,column=1).border=TB
        ws.cell(row=r,column=2,value=a['spec']).font=DF; ws.cell(row=r,column=2).border=TB
        for c,e in enumerate(['Al','Cu','Zn','Mg','Mn','Ag','Li','Ni','Fe','Si','Zr','Ti'],3):
            v=a['comp'].get(e,0); cl=ws.cell(row=r,column=c,value=v if v>0 else '-'); cl.border=TB
            cl.font=(RF if e=='Ag' else (GF if e=='Li' else DF)) if v>0 else Font(name='Arial',size=10,color='BBBBBB')
            if v>0: cl.number_format='0.00'
    ws.column_dimensions['A'].width=12; ws.column_dimensions['B'].width=12
    for c in range(3,15): ws.column_dimensions[get_column_letter(c)].width=8
    # Stage sheets
    _stage_sheet(wb,"A) Raw Material Cost","A) Raw Material Cost (USD/kg)","Element cost per kg — before conversion",df_hist,alloys,ak,rb,rt,'raw',SFill['raw'])
    _stage_sheet(wb,"B) Billet Cost",f"B) Billet Cost (USD/kg) — ×{rb:.4f}",f"Yield {1/rb*100:.1f}% | No scrap credits",df_hist,alloys,ak,rb,rt,'billet',SFill['billet'])
    _stage_sheet(wb,"C) Extr.-Forging Cost",f"C) Extrusion / Forging Cost (USD/kg) — ×{rt:.4f}",
        f"×{rb:.4f} (ingot→billet) × {re_:.1f} (billet→extr./forg.) = ×{rt:.4f} | Yield {1/rt*100:.1f}%",
        df_hist,alloys,ak,rb,rt,'ext',SFill['ext'],extra=True)
    # All stages combined
    ws5=wb.create_sheet("All Stages Combined"); ws5.merge_cells('A1:Q1')
    ws5['A1'].value='Material Cost per kg at Each Stage (USD/kg)'; ws5['A1'].font=Font(name='Arial',bold=True,size=14,color='2F5496')
    ws5.merge_cells('A2:Q2'); ws5['A2'].value=f'Raw → Billet (×{rb:.3f}) → Extr./Forg. (×{rt:.3f}) | No scrap credits'; ws5['A2'].font=SF
    for cs,lbl,stg in [(2,'RAW ($/kg)','raw'),(7,f'BILLET ($/kg) ×{rb:.3f}','billet'),(12,f'EXTR./FORG. ($/kg) ×{rt:.3f}','ext')]:
        ws5.merge_cells(start_row=3,start_column=cs,end_row=3,end_column=cs+4)
        cl=ws5.cell(row=3,column=cs,value=lbl); cl.font=Font(name='Arial',bold=True,size=10,color='FFFFFF'); cl.fill=SH[stg]; cl.alignment=Alignment(horizontal='center'); cl.border=TB
        for cc in range(cs+1,cs+5): ws5.cell(row=3,column=cc).fill=SH[stg]; ws5.cell(row=3,column=cc).border=TB
    ws5.cell(row=3,column=1).border=TB; cl=ws5.cell(row=4,column=1,value='Date'); cl.font=HF; cl.fill=HFL; cl.alignment=Alignment(horizontal='center'); cl.border=TB
    for so in [1,6,11]:
        for j,key in enumerate(ak):
            cl=ws5.cell(row=4,column=so+1+j,value=alloys[key]['name']); cl.font=Font(name='Arial',bold=True,size=9,color='FFFFFF')
            cl.fill=PatternFill('solid',fgColor=AC[key]); cl.alignment=Alignment(horizontal='center'); cl.border=TB
    for i in range(N):
        row=df_hist.iloc[i]; r=5+i; last=(i==N-1); lbl=row['date']+(' ← LATEST' if last else '')
        cl=ws5.cell(row=r,column=1,value=lbl); cl.font=RF if last else DF; cl.border=TB
        if last: cl.fill=TODAY_FILL
        hp=_hp(row)
        for j,key in enumerate(ak):
            raw,_,_=calc_alloy_cost(alloys[key]['comp'],hp); bil,ext=calc_conversion_costs(raw,rb,rt)
            for co,val,stg in [(1,raw,'raw'),(6,bil,'billet'),(11,ext,'ext')]:
                cl=ws5.cell(row=r,column=co+1+j,value=round(val,2))
                cl.font=Font(name='Arial',bold=last or stg=='ext',size=10 if stg=='ext' else 9,color=AC[key])
                cl.fill=TODAY_FILL if last else SFill[stg]; cl.border=TB; cl.number_format='0.00'
    ws5.column_dimensions['A'].width=20
    for c in range(2,17): ws5.column_dimensions[get_column_letter(c)].width=10
    # Sources
    ws6=wb.create_sheet("Sources & Methodology"); ws6.merge_cells('A1:C1')
    ws6['A1'].value='Data Sources & Methodology'; ws6['A1'].font=Font(name='Arial',bold=True,size=14,color='2F5496')
    for c,h in enumerate(['Element','Unit','Source'],1):
        cl=ws6.cell(row=3,column=c,value=h); cl.font=HF; cl.fill=HFL; cl.alignment=Alignment(horizontal='center'); cl.border=TB
    for i,(e,u,s) in enumerate([('Aluminium','USD/t','Westmetall LME Cash'),('Copper','USD/t','Westmetall LME Cash'),
        ('Silver','USD/oz','goldprice.org / Bullion.com / APMEX / JM Bullion / Fortune'),('Zinc','USD/t','Westmetall LME Cash'),
        ('Nickel','USD/t','Westmetall LME Cash'),('Lithium','USD/kg','TradingEcon Li₂CO₃ ×10 / ChemAnalyst / IMARC'),
        ('Magnesium','USD/kg','TradingEcon / Asian Metal'),('Manganese','USD/kg','TradingEcon / Asian Metal'),
        ('Titanium','USD/kg','TradingEcon / Asian Metal'),('Silicon','USD/kg','TradingEcon / Asian Metal'),
        ('Zirconium','USD/kg','USGS ($35/kg)'),('Iron','USD/kg','Nominal ($0.10/kg)')]):
        r=4+i; ws6.cell(row=r,column=1,value=e).font=BF; ws6.cell(row=r,column=1).border=TB
        ws6.cell(row=r,column=2,value=u).font=DF; ws6.cell(row=r,column=2).border=TB
        ws6.cell(row=r,column=3,value=s).font=DF; ws6.cell(row=r,column=3).border=TB
    r=18; ws6.cell(row=r,column=1,value="CONVERSION:").font=RF
    for i,t in enumerate([f"Stage 1: Ingot→Billet = ×{rb:.4f} (yield {1/rb*100:.1f}%)",
        f"Stage 2: Billet→Extr./Forg. = ×{re_:.1f} (yield {1/re_*100:.1f}%)",
        f"Total: ×{rt:.4f} (yield {1/rt*100:.1f}%)","","Raw material only — excludes processing, handling, logistics.","No scrap value credited."]):
        ws6.cell(row=r+1+i,column=1,value=t).font=DF
    ws6.column_dimensions['A'].width=70; ws6.column_dimensions['B'].width=10; ws6.column_dimensions['C'].width=60
    buf=BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()
